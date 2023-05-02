from email import message
import time
import pandas as pd
import numpy as np
import re
import datetime
import csv
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
import json
import seaborn as sns
import matplotlib.pyplot as plt
import psycopg2
import asyncio
import logging
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram import Bot, Dispatcher, executor, types
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
import openpyxl
import io
from dotenv import load_dotenv
import os

load_dotenv()


logging.basicConfig(level=logging.INFO)
bot = Bot(os.getenv('TOKEN'))
dp = Dispatcher(bot, storage=MemoryStorage())

A = []
users_random_message = []
ADV = []
info_user_message = []
ADDmes = []

class Statess(StatesGroup):
    #Состояние для кнопки помощи пользователям
    help_button = State()
    #Состояния на добавление результатов
    user_enter = State()
    category_enter = State()
    note_ctg_enter = State()
    zp_rs = State()
    size_s = State()
    card_cash = State()
    #Состояния на советы
    advices = State()
    #Состояние на вывод информации через бота
    info_enter = State()
    #Состояния для редактирвоания данных
    red_set = State()
    red_enter = State()
    red_1 = State()
    red_2 = State()
    red_3 = State()
    red_4 = State()
    red_5 = State()
    #Состояние для передачи эксель файла
    xls_set = State() 

@dp.message_handler(commands="start") # Начало работы с ботом
async def startmessage(message: types.Message):
    markupaa = types.InlineKeyboardMarkup(row_width=2)
    btn = types.InlineKeyboardButton("Help", callback_data="helpmsg")
    markupaa.add(btn)
    await bot.send_message(message.chat.id, "Привет, я - финансовый бот, помогу тебе оптимизировать твои расходы. Для начала работы нажми 'Help'  ", reply_markup=markupaa)
    try:
        user = str(message.from_user.id)
        print(user)
        # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', # ноутбук
        #                 password='KokoRari-23', host='localhost', port='5432')
        db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
                password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))
        cur = db.cursor()

        cur.execute(f'''CREATE TABLE "{user}" (
            id SERIAL NOT NULL PRIMARY KEY,
            Категория TEXT,
            Примечания TEXT,
            ЗП_РС TEXT,
            Размер INTEGER DEFAULT 0,
            Карта_Нал TEXT,
            time date NOT NULL
            )''')
        
        db.commit()
        db.close()
    except psycopg2.errors.DuplicateTable:
        pass

        

@dp.callback_query_handler(text="helpmsg")
async def callback(call: types.CallbackQuery):
            markup_commands = types.InlineKeyboardMarkup(row_width=3)
            crt = types.InlineKeyboardButton("Редактировать", callback_data="redact")
            add = types.InlineKeyboardButton("Добавить", callback_data="add")
            gdb = types.InlineKeyboardButton("Выслать", callback_data="db")
            inf = types.InlineKeyboardButton("Информация", callback_data="info")
            anl = types.InlineKeyboardButton("Анализ", callback_data="analize")
            adv = types.InlineKeyboardButton("Советы", callback_data="advice")
            markup_commands.add(crt, add, gdb, inf, anl, adv)
            
            await bot.send_message(call.message.chat.id, '''Мои функции:
    🔧Редактировать - Удаление всей базы данных или перезапись последних веденных данных
    📝Добавить - добавление доходов/расходов
    📤Выслать - высылается excel файл с данными 
    📖Информация - отображение информации, содержащейся в базе данных
    📊Анализ - Визуализация данных и показ наиболее часто совершаемых процедур/избранные категории 
    💡Советы - советы по оптимизации доходов и расходов''', reply_markup=markup_commands)

@dp.callback_query_handler(text="redact")
async def redact_ex(message: types.Message):
    markup_redact = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True, one_time_keyboard=True)
    red_del = types.InlineKeyboardButton('Удалить данные')
    red_last = types.InlineKeyboardButton('Перезаписать последние данные')
    red_excel = types.InlineKeyboardButton('Внести Excel')
    markup_redact.add(red_del, red_last, red_excel)
    await bot.send_message(message.from_user.id, '''✍️*В редактировании можно выполнить несколько команд:*
❌Удалить данные - удаляются все данные
📝Перезаписать последние данные - последние данные, которые вы ввели можно изменить, если допустили ошибку
📩Внести Excel - вытекающее из функции "Выслать", если вы самостоятельно редатировали xlsx, xsl файлы, их можно внести обратно''', parse_mode='Markdown', reply_markup=markup_redact)
    await Statess.red_set.set()

@dp.message_handler(state=Statess.red_set)
async def red_del_func(message: types.Message, state: FSMContext):
    # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', # ноутбук
    #         password='KokoRari-23', host='localhost', port='5432')
    db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
            password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))
    cur = db.cursor()
    async with state.proxy() as delete_dt:
        user_red = str(message.from_user.id)
        print(user_red)
        if message.text == 'Удалить данные':
            markup_del_red = types.InlineKeyboardMarkup(row_width=2)
            btn_del_red = types.InlineKeyboardButton("Вернуться в главное меню", callback_data="helpmsg")
            markup_del_red.add(btn_del_red)
            cur.execute(f'''DELETE FROM "{user_red}"''')
            db.commit()
            db.close()
            await bot.send_message(message.from_user.id, '''Данные были полностью удалены из базы!''', reply_markup=markup_del_red)
            await state.finish()
        if message.text == 'Перезаписать последние данные':
            markup_vvod = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True, one_time_keyboard=True)
            btn_vvod = types.InlineKeyboardButton('Ввод')
            markup_vvod.add(btn_vvod)
            await bot.send_message(message.from_user.id, '''Введите любое сообщение для подтверждения''', reply_markup=markup_vvod)
            await Statess.red_enter.set()
        if message.text == 'Внести Excel':
            await bot.send_message(message.from_user.id,'''Внесите Excel файл:''')
            await Statess.xls_set.set()

@dp.message_handler(state=Statess.xls_set, content_types=['document'])
async def xls_import(message: types.Message, state: FSMContext):
            async with state.proxy() as xls_file:
                markup_del_xls = types.InlineKeyboardMarkup(row_width=2)
                btn_del_xls = types.InlineKeyboardButton("Вернуться в главное меню", callback_data="helpmsg")
                markup_del_xls.add(btn_del_xls)
                # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', # ноутбук
                #         password='KokoRari-23', host='localhost', port='5432')
                db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', #Комп
                        password='Ghg3500pol', host='localhost', port='5432')
                cur = db.cursor()
                user_xls = str(message.from_user.id)
                file_path = await bot.download_file_by_id(message.document.file_id)
                wb = openpyxl.load_workbook(filename=file_path)
                ws = wb.active
                rows_xls = ws.rows
                cur.execute(f'''SELECT * from "{user_xls}" ''')
                res_dt_postgre = cur.fetchall()
                print(res_dt_postgre)
                cur.execute(f'''DELETE FROM "{user_xls}" ''')
                for row_xls in rows_xls:
                    values_xls = [cell.value for cell in  row_xls]
                    print(values_xls[0], "авыа")
                    cur.execute(f'''INSERT INTO "{user_xls}" (id, Категория, Примечания, ЗП_РС, Размер, Карта_Нал, time) VALUES(%s, %s, %s, %s, %s, %s, %s)''', (values_xls[0], values_xls[1], values_xls[2], values_xls[3], values_xls[4], values_xls[5], values_xls[6]))

                await bot.send_message(message.from_user.id, '''Данные из Excel файла были успешно внесены в базу!''', reply_markup=markup_del_xls)
                db.commit() 
                db.close()
                await state.finish()

@dp.message_handler(state=Statess.red_enter)
async def red_ctg(message: types.Message, state: FSMContext):
        async with state.proxy() as ctg:
            markup = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True, one_time_keyboard=True)
            item = types.InlineKeyboardButton('Зарплата')
            item_freetime = types.InlineKeyboardButton('Развлечения')
            item_food = types.InlineKeyboardButton('Продукты')
            item_fastfood = types.InlineKeyboardButton('Рестораны')
            item_transport = types.InlineKeyboardButton('Транспорт')
            item_hoz = types.InlineKeyboardButton('Хозтовары')
            item_sub = types.InlineKeyboardButton('Подписки')
            item_health = types.InlineKeyboardButton('Здоровье')
            item_extrem = types.InlineKeyboardButton('Экстренные расходы')
            markup.add(item, item_freetime, item_food, item_fastfood, item_transport, item_hoz, item_sub, item_health, item_extrem)
            await bot.send_message(message.chat.id, "Введите категорию: ", reply_markup=markup)
            ctg['category_enter'] = message.text
            users_random_message.append(message.text)

            # print(A)
            await Statess.next()

@dp.message_handler(state=Statess.red_1)    
async def red_notee(message: types.Message, state: FSMContext):
    async with state.proxy() as note:   
        await bot.send_message(message.from_user.id, "Введите примечания к категории: ")
        note['note_ctg_enter'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.red_2)   
async def red_sal(message: types.Message, state: FSMContext):
    async with state.proxy() as z_s:
        markup_change = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        item_change = types.InlineKeyboardButton('Доход')
        item_change2 = types.InlineKeyboardButton('Расход')
        markup_change.add(item_change, item_change2)
        await bot.send_message(message.from_user.id, "Укажите вид(Доход или Расход): ", reply_markup=markup_change)
        z_s['zp_rs'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.red_3) 
async def red_sizee(message: types.Message, state: FSMContext):
    async with state.proxy() as se:    
        await bot.send_message(message.from_user.id, "Введите размер затрат/дохода: ")
        se['size_s'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.red_4)     
async def red_card_or_cash(message: types.Message, state: FSMContext):
    async with state.proxy() as cc:
        markup_ccn = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
        item_card = types.InlineKeyboardButton('Карта')
        item_cash = types.InlineKeyboardButton('Наличные')
        item_qrcode = types.InlineKeyboardButton('Qr-code')
        item_NFC = types.InlineKeyboardButton('NFC')    
        markup_ccn.add(item_card, item_cash, item_NFC, item_qrcode)
        await bot.send_message(message.from_user.id, "Введите способ получения/оплаты: ", reply_markup=markup_ccn)
        cc['card_cash'] = message.text
        A.append(message.text)
        print(A)  
        await Statess.next() 

@dp.message_handler(state=Statess.red_5)    
async def red_update_db(message: types.Message, state: FSMContext):
    A.append(message.text)
    print(A)
    # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres',  #Ноут
    #                     password='KokoRari-23', host='localhost', port='5432')

    db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
            password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))

    cur = db.cursor()
    user_add = str(message.from_user.id)
    cur.execute(f'''SELECT * FROM "{user_add}" ORDER BY id DESC LIMIT 1''')
    last_dt = cur.fetchone()
    cur.execute(f'''UPDATE "{user_add}" SET Категория = %s, Примечания = %s, ЗП_РС = %s, Размер = %s, Карта_Нал = %s WHERE id = %s''', [A[0], A[1], A[2], A[3], A[4], last_dt[0]])
    db.commit()
    db.close()
    A.clear()
    print(A)
    delete = types.ReplyKeyboardRemove()
    markupADD = types.InlineKeyboardMarkup(row_width=2)
    btn3  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
    markupADD.add(btn3)
    await bot.send_message(message.from_user.id, "Данные были успешно обновлены!", reply_markup=delete)
    await bot.send_message(message.from_user.id, 'Для продолжения работы нажми "Help"', reply_markup=markupADD)
    await state.finish()
        
@dp.callback_query_handler(text="add")
async def addexp_ctg(call: types.CallbackQuery):
    markup_help_users = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True, one_time_keyboard=True)
    item_help = types.InlineKeyboardButton('Инструкция')
    item_enter = types.InlineKeyboardButton('Ввод')
    item_back = types.InlineKeyboardButton('Отмена', callback_data="helpmsg")
    markup_help_users.add(item_help, item_enter, item_back)
    await bot.send_message(call.message.chat.id, '''Выберите дальнейшее действие: ''', reply_markup=markup_help_users)
    await Statess.help_button.set()

@dp.message_handler(state=Statess.help_button)
async def help_function(message: types.Message, state: FSMContext):
    async with state.proxy() as instruction:
        if message.text == 'Инструкция':
            markup_help_users_2 = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
            item_enter = types.InlineKeyboardButton('Ввод')
            # otmena = types.InlineKeyboardButton("Отмена", callback_data="helpmsg")
            markup_help_users_2.add(item_enter)
            await bot.send_message(message.from_user.id,'''
*КОРРЕКТНОЕ ВВЕДЕНИЕ ДАННЫХ В БАЗУ*
📚*Категория* - ввод определенного вида дохода или расхода.
Примеры: еда, транспорт, премия, зарплата и т.д.
📑*Примечания* - ввод дополнительной информации к категории
Примеры: ресторан, метро, проект. 
*В случае получения ЗП можно писать ДОХОД*
📋*Зарплата/Расходы* - текстовое пояснение вида категории(Либо ДОХОД, Либо РАСХОД)
💵*Размер* - сумма затрат или дохода
💳*Карта/Наличные* - способ получения/оплаты(для более детальной визуализации можно вводить не только Карта или Наличные, но и NFC и т.д.)

*ВАЖНО!!!*
Данные вводятся поштучно. То есть ввести сразу все не получится
*ПОСЛЕ ПРОЧТЕНИЯ ИНСТРУКЦИИ ВВЕДИТЕ ЛЮБОЙ СИМВОЛ ИЛИ НАЖМИТЕ НА КНОПКУ "Ввод" ДЛЯ НАЧАЛА РАБОТЫ*
        ''', parse_mode='Markdown', reply_markup=markup_help_users_2)
            await Statess.user_enter.set()
        if message.text == 'Отмена':
            await state.finish()
            markup_back = types.InlineKeyboardMarkup(row_width=2)
            btnBK  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
            markup_back.add(btnBK)
            await bot.send_message(message.from_user.id, 'Нажмите "Help", чтобы вернуться к главное меню', reply_markup=markup_back)
        if message.text == 'Ввод': 
            await bot.send_message(message.from_user.id, '''Напишите любое сообщение для подтверждения''')
            await Statess.user_enter.set()
        if message.text != 'Отмена' and message.text != 'Ввод' and message.text != 'Инструкция':
            await bot.send_message(message.from_user.id, '''Напишите любое сообщение для подтверждения''')
            await Statess.user_enter.set()
#         else:
            # markup_back_to_menu = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
            # mbtm = types.InlineKeyboardButton("❌Отмена", callback_data="helpmsg")
            # markup_back_to_menu.add(mbtm)
#             await bot.send_message(message.from_user.id, '''Для подтверждения действия напишите любое сообщение или символ
# Для отмены нажмите на кнопку "❌Отмена" ''', reply_markup=markup_back_to_menu)
#             await Statess.user_enter.set()


# dp.message_handler(content_types=types.ContentTypes.TEXT, state=Statess.help_button)
# async def next_add(message: types.Message):
#      if message.text == 'Ввод':
#             await Statess.next()
        
@dp.message_handler(state=Statess.user_enter)
async def addexp_ctg(message: types.Message, state: FSMContext):
        async with state.proxy() as ctg:
            markup = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True, one_time_keyboard=True)
            item = types.InlineKeyboardButton('Зарплата')
            item_freetime = types.InlineKeyboardButton('Развлечения')
            item_food = types.InlineKeyboardButton('Продукты')
            item_fastfood = types.InlineKeyboardButton('Рестораны')
            item_transport = types.InlineKeyboardButton('Транспорт')
            item_hoz = types.InlineKeyboardButton('Хозтовары')
            item_sub = types.InlineKeyboardButton('Подписки')
            item_health = types.InlineKeyboardButton('Здоровье')
            item_extrem = types.InlineKeyboardButton('Экстренные расходы')
            markup.add(item, item_freetime, item_food, item_fastfood, item_transport, item_hoz, item_sub, item_health, item_extrem)
            await bot.send_message(message.chat.id, "Введите категорию: ", reply_markup=markup)
            ctg['category_enter'] = message.text
            users_random_message.append(message.text)

            # print(A)
            await Statess.next()

@dp.message_handler(state=Statess.category_enter)    
async def addexp_notee(message: types.Message, state: FSMContext):
    async with state.proxy() as note:   
        await bot.send_message(message.from_user.id, "Введите примечания к категории: ")
        note['note_ctg_enter'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.note_ctg_enter)   
async def addexp_sal(message: types.Message, state: FSMContext):
    async with state.proxy() as z_s:
        markup_change = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        item_change = types.InlineKeyboardButton('Доход')
        item_change2 = types.InlineKeyboardButton('Расход')
        markup_change.add(item_change, item_change2)
        await bot.send_message(message.from_user.id, "Укажите вид(Доход или Расход): ", reply_markup=markup_change)
        z_s['zp_rs'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.zp_rs) 
async def addexp_sizee(message: types.Message, state: FSMContext):
    async with state.proxy() as se:    
        await bot.send_message(message.from_user.id, "Введите размер затрат/дохода: ")
        se['size_s'] = message.text
        A.append(message.text)
        print(A)
        await Statess.next()

@dp.message_handler(state=Statess.size_s)     
async def addexp_card_or_cash(message: types.Message, state: FSMContext):
    async with state.proxy() as cc:
        markup_ccn = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
        item_card = types.InlineKeyboardButton('Карта')
        item_cash = types.InlineKeyboardButton('Наличные')
        item_qrcode = types.InlineKeyboardButton('Qr-code')
        item_NFC = types.InlineKeyboardButton('NFC')    
        markup_ccn.add(item_card, item_cash, item_NFC, item_qrcode)
        await bot.send_message(message.from_user.id, "Введите способ получения/оплаты: ", reply_markup=markup_ccn)
        cc['card_cash'] = message.text
        A.append(message.text)
        print(A)  
        await Statess.next() 

@dp.message_handler(state=Statess.card_cash)    
async def save_db(message: types.Message, state: FSMContext):
    A.append(message.text)
    print(A)
    curData = datetime.date.today()
    # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres',  #Ноут
    #                     password='KokoRari-23', host='localhost', port='5432')

    db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
            password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))

    cur = db.cursor()
    user_add = str(message.from_user.id)
    cur.execute(f'''INSERT INTO "{user_add}" (Категория, Примечания, ЗП_РС, Размер, Карта_Нал, time) VALUES(%s, %s, %s, %s, %s, %s)''', [A[0], A[1], A[2], A[3], A[4], curData])
    db.commit()
    db.close()
    A.clear()
    print(A)
    delete = types.ReplyKeyboardRemove()
    markupADD = types.InlineKeyboardMarkup(row_width=2)
    btn3  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
    markupADD.add(btn3)
    await bot.send_message(message.from_user.id, "Данные были успешно добавлены!", reply_markup=delete)
    await bot.send_message(message.from_user.id, 'Для продолжения работы нажми "Help"', reply_markup=markupADD)
    await state.finish()

@dp.callback_query_handler(text= "analize")
async def analize(message):
    try:
        # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', #Ноут
        #                     password='KokoRari-23', host='localhost', port='5432')

        db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
                password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))

        cur = db.cursor()
        user_analize = str(message.from_user.id)
        cur.execute(f'''SELECT * FROM "{user_analize}" ''')
        analize_res = cur.fetchall()
        print(analize_res)
        headers = [col[0] for col in cur.description] # get headers
        analize_res.insert(0, tuple(headers))
        fp = open('fin_tab.csv', 'w', encoding="utf-8", newline = '') 
        myFile = csv.writer(fp)
        myFile.writerows(analize_res)
        fp.close()

        data = pd.read_csv('fin_tab.csv')
        data_copy = data.copy()
        # print(data_copy)
        df = data_copy.drop('id', axis=1)
        df2 = df.loc[df["Категория"] != "Зарплата"]
        df_sum = df.loc[df["Категория"] == "Зарплата", "Размер"].sum()
        df_sum_exp = df.loc[df["ЗП_РС"] == "Расход", "Размер"].sum()
        copy = df2.copy()
        print(copy)
        print(df_sum)
        print(df_sum_exp)

        plt.subplots()
        sns.catplot(data=copy, x="Категория", y="Размер", kind="bar", ci=None)
        plt.tick_params('x', rotation=90)
        
        plt.title("Анализ данных")
        plt.tight_layout()
        plt.savefig('analize_data.png')

        await bot.send_photo(message.from_user.id, photo=open('analize_data.png', 'rb'))
        await bot.send_message(message.from_user.id,f'''Выше показан анализ ваших затрат на различные категории за все время.
    *Ваш доход - {df_sum}*
    *Ваши расходы - {df_sum_exp}*''', parse_mode='Markdown')
        

        cur.execute(f'''SELECT * FROM "{user_analize}" WHERE date_trunc('month', time) = date_trunc('month', current_timestamp) ''')
        analize_res_month = cur.fetchall()
        print(analize_res_month)
        headers_month = [col_month[0] for col_month in cur.description] # get headers
        analize_res_month.insert(0, tuple(headers_month))
        fpm = open('fin_tab_Month.csv', 'w', encoding="utf-8", newline = '') 
        myFileM = csv.writer(fpm)
        myFileM.writerows(analize_res_month)
        fpm.close()
        data_grM = pd.read_csv('fin_tab_Month.csv')
        data_grcM = data_grM.copy()
        # print(data_copy)
        df_grM = data_grcM.drop('id', axis=1)
        df_gr_2M = df_grM.loc[df_grM["Категория"] != "Зарплата"]
        df_gr_sum_M = df_grM.loc[df_grM["Категория"] == "Зарплата", "Размер"].sum()
        df_gr_sum_exp_M = df_grM.loc[df_grM["ЗП_РС"] == "Расход", "Размер"].sum()
        print(df_gr_sum_M)
        print(df_gr_sum_exp_M)
        copy_grM = df_gr_2M.copy()
        print(copy_grM)
        
        plt.subplots()
        sns.catplot(data=copy_grM, x="Категория", y="Размер", kind="bar", ci=None)
        plt.tick_params('x', rotation=90)

        plt.title("Анализ данных")
        plt.tight_layout()
        plt.savefig('analize_data_month.png')

        await bot.send_photo(message.from_user.id, photo=open('analize_data_month.png', 'rb'))
        await bot.send_message(message.from_user.id,f'''На данной картинке вы можете внимательно изучить свои затраты на месяц.
    *Ваш доход за месяц- {df_gr_sum_M}*
    *Ваши расходы за месяц- {df_gr_sum_exp_M}*''', parse_mode='Markdown')
        
        cur.execute(f'''SELECT * FROM "{user_analize}" WHERE date_trunc('week', time) = date_trunc('week', current_timestamp) ''')
        analize_res_week = cur.fetchall()
        print(analize_res_week)
        headers_week = [col_week[0] for col_week in cur.description] # get headers
        analize_res_week.insert(0, tuple(headers_week))
        fpw = open('fin_tab_week.csv', 'w', encoding="utf-8", newline = '') 
        myFileW = csv.writer(fpw)
        myFileW.writerows(analize_res_week)
        fpw.close()
        data_gr_week = pd.read_csv('fin_tab_week.csv')
        data_grc_week = data_gr_week.copy()
        # print(data_copy)
        df_grW = data_grc_week.drop('id', axis=1)
        df_gr_2W = df_grW.loc[df_grW["Категория"] != "Зарплата"]
        df_gr_sum_W = df_grW.loc[df_grW["Категория"] == "Зарплата", "Размер"].sum()
        df_gr_sum_exp_W = df_grW.loc[df_grW["ЗП_РС"] == "Расход", "Размер"].sum()
        print(df_gr_sum_W)
        print(df_gr_sum_exp_W)
        copy_gr_week = df_gr_2W.copy()
        print(copy_gr_week)

        plt.subplots()
        sns.catplot(data=copy_gr_week, x="Категория", y="Размер", kind="bar", ci=None)
        plt.tick_params('x', rotation=90)

        plt.title("Анализ данных")
        plt.tight_layout()
        plt.savefig('analize_data_week.png')

        await bot.send_photo(message.from_user.id, photo=open('analize_data_week.png', 'rb'))
        await bot.send_message(message.from_user.id,f'''Здесь проведен анализ наиболее недавних сделанных вами транзакций(За неделю). Внимательно изучите каждую категорию и сделайте выводы.
    *Ваш доход за неделю- {df_gr_sum_W}*
    *Ваши расходы за неделю- {df_gr_sum_exp_W}*''', parse_mode='Markdown')
        
        markup_analytics = types.InlineKeyboardMarkup(row_width=2)
        btnAn  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
        markup_analytics.add(btnAn)
        await bot.send_message(message.from_user.id, 'Нажмите "Help", чтобы вернуться к главное меню', reply_markup=markup_analytics)
    except ValueError:
        await bot.send_message(message.from_user.id, '''Для аналитики за неделю требуется больше данных''')
        markup_analytics_2 = types.InlineKeyboardMarkup(row_width=2)
        btnAn_2  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
        markup_analytics_2.add(btnAn_2)
        await bot.send_message(message.from_user.id, 'Нажмите "Help", чтобы вернуться к главное меню', reply_markup=markup_analytics_2)
         
@dp.callback_query_handler(text= "advice")
async def advice_list(message):
    markup_advice = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True, one_time_keyboard=True)
    invest = types.InlineKeyboardButton('Инвестиции')
    jkh = types.InlineKeyboardButton('Дом')
    # prioritets = types.InlineKeyboardButton('Приоритеты')
    markup_advice.add(invest, jkh)
    await bot.send_message(message.from_user.id, '''
Список советов:
Инвестиции - больше предназначены для начинающих инвесторов
Дом - советы для распределения финансов в доме
''', parse_mode='Markdown', reply_markup=markup_advice)
    await Statess.advices.set()

from commands import *
import random
@dp.message_handler(state=Statess.advices)  
async def change_adv(message: types.Message, state: FSMContext):
        ADV.append(message.text)
        print(ADV)
        if ADV[0] == "Инвестиции":
            a = random.choice(investition)
            await bot.send_message(message.chat.id, f'{a}')
        if ADV[0] == "Дом":
            b = random.choice(zhkh)
            await bot.send_message(message.chat.id, f'{b}')
        
        markupADV = types.InlineKeyboardMarkup(row_width=2)
        btnAD  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
        markupADV.add(btnAD)
        ADV.clear()
        await state.finish()
        await bot.send_message(message.from_user.id, 'Нажмите "Help", чтобы вернуться к главное меню', reply_markup=markupADV)

@dp.callback_query_handler(text="db") 
async def file_table(message):
    # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', #Ноут
    #                     password='KokoRari-23', host='localhost', port='5432')
    
    db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
            password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))

    cur = db.cursor()
    user_table = str(message.from_user.id)

    book = openpyxl.Workbook()
    sheet = book.active

    cur.execute(f'''SELECT * FROM "{user_table}" ''')
    results_db = cur.fetchall()
    print(results_db)
    i = 0
    for row_db in results_db:
        i += 1
        j = 1
        for col in row_db:
            cell = sheet.cell(row = i, column = j)
            cell.value = col
            j += 1
    
    # (D) SAVE EXCEL FILE & CLOSE DB
    book.save("Финансы.xlsx")
    db.close()
    await bot.send_message(message.from_user.id, '''Ваша База данных, пожалуйста:''')
    await bot.send_document(message.from_user.id, open(r'Финансы.xlsx', 'rb'))
    markupBASE = types.InlineKeyboardMarkup(row_width=2)
    btn4  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
    markupBASE.add(btn4)
    await bot.send_message(message.from_user.id, 'Нажмите "Help", чтобы вернуться к главное меню', reply_markup=markupBASE)

@dp.callback_query_handler(text="info")
async def info(message):
    markup_info = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
    btn_day = types.InlineKeyboardButton("За день")
    btn_month = types.InlineKeyboardButton("За неделю")
    btn_year = types.InlineKeyboardButton("За месяц")
    btn_all = types.InlineKeyboardButton("За год")
    markup_info.add(btn_day, btn_month, btn_year, btn_all)
    await bot.send_message(message.from_user.id, '''
За какой промежуток времени вам вывести информацию?
1 - за день
2 - за неделю
3 - за месяц
4 - за год
''', parse_mode='Markdown', reply_markup=markup_info)
    await Statess.info_enter.set()

@dp.message_handler(state=Statess.info_enter)
async def info_day(message: types.Message, state: FSMContext):
    async with state.proxy() as inf:
        inf['info_user_message'] = message.text
        info_user_message.append(message.text)
        print(info_user_message)
        delete_day = types.ReplyKeyboardRemove() 
        # db =  psycopg2.connect(dbname='TelegramFinanseBot', user='postgres', 
        #                 password='KokoRari-23', host='localhost', port='5432')

        db =  psycopg2.connect(dbname=(os.getenv('DBNAME')), user=(os.getenv('USER')), #Комп
                password=(os.getenv('PASSWORD')), host=(os.getenv('HOST')), port=(os.getenv('PORT')))
        cur = db.cursor()
        user_view = str(message.from_user.id)
        date = datetime.date.today()
        print(date)
        if info_user_message[0] == "За день":
            print(date)
            cur.execute(f'''SELECT Категория, Примечания, ЗП_РС, Размер, Карта_Нал, TO_DATE(to_char(time, 'YYYY-MM-DD'), 'YYYY-MM-DD') as time FROM "{user_view}" WHERE time = to_timestamp('{date}', 'YYYY-MM-DD') ''')
            data_result = cur.fetchall()
            print(data_result)
            if data_result == []:
                await bot.send_message(message.from_user.id, '''За сегодняшний день вы еще не внесли никаких данных''')
            else:
                for row in data_result:
                    await bot.send_message(message.from_user.id, f'''<pre>Категория: {row[0]} 
Примечания: {row[1]} 
Зарплата/расход: {row[2]} 
Размер: {row[3]} 
Карта/Наличные: {row[4]} 
Время: {row[5]}</pre>''', parse_mode="HTML")
                    
        if info_user_message[0] == "За месяц":
            date_month = datetime.date.today()
            month = date_month.strftime("%Y-%m")
            print(month)
            cur.execute(f'''SELECT Категория, Примечания, ЗП_РС, Размер, Карта_Нал, time FROM "{user_view}" WHERE date_trunc('month', time) = date_trunc('month', current_timestamp) ''')
            month_result = cur.fetchall()
            print(month_result)
            if month_result == []:
                await bot.send_message(message.from_user.id, '''За текущий месяц вы еще не внесли никаких данных''')
            else:
                for row_month in month_result:
                        await bot.send_message(message.from_user.id, f'''<pre>Категория: {row_month[0]} 
Примечания: {row_month[1]} 
Зарплата/расход: {row_month[2]} 
Размер: {row_month[3]} 
Карта/Наличные: {row_month[4]} 
Время: {row_month[5]}</pre>''', parse_mode="HTML")
                        
        if info_user_message[0] == "За год":
            date_year = datetime.date.today()
            year = date_year.strftime("%Y")
            print(year)
            cur.execute(f'''SELECT Категория, Примечания, ЗП_РС, Размер, Карта_Нал, time FROM "{user_view}" WHERE date_trunc('year', time) = date_trunc('year', current_timestamp) ''')
            year_result = cur.fetchall()
            print(year_result)
            if year_result == []:
                await bot.send_message(message.from_user.id, '''За текущий год вы еще не внесли никаких данных''')
            else:
                for row_year in year_result:
                        await bot.send_message(message.from_user.id, f'''<pre>Категория: {row_year[0]} 
Примечания: {row_year[1]} 
Зарплата/расход: {row_year[2]} 
Размер: {row_year[3]} 
Карта/Наличные: {row_year[4]} 
Время: {row_year[5]}</pre>''', parse_mode="HTML")

        if info_user_message[0] == "За неделю":
            cur.execute(f'''SELECT Категория, Примечания, ЗП_РС, Размер, Карта_Нал, time FROM "{user_view}" WHERE date_trunc('week', time) = date_trunc('week', current_timestamp) ''')
            res_all = cur.fetchall()
            print(res_all)
            if res_all == []:
                await bot.send_message(message.from_user.id, '''За эту неделю еще не было внесено никаких данных''')
            else:
                for row in res_all:
                        await bot.send_message(message.from_user.id, f'''<pre>Категория: {row[0]} 
Примечания: {row[1]} 
Зарплата/расход: {row[2]} 
Размер: {row[3]} 
Карта/Наличные: {row[4]} 
Время: {row[5]}</pre>''', parse_mode="HTML")
                        
        info_user_message.clear()
        markupINFO = types.InlineKeyboardMarkup(row_width=2)
        btn4  = types.InlineKeyboardButton("Help", callback_data="helpmsg")
        markupINFO.add(btn4)
        await bot.send_message(message.from_user.id, 'Для продолжения работы нажми "Help"', reply_markup=markupINFO)
        db.commit()
        db.close() 
    await state.finish()  

@dp.message_handler(content_types=['text'])
async def get_text_messages(message):
    markup_start = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True, one_time_keyboard=True)
    start = types.InlineKeyboardButton("/start")
    markup_start.add(start)
    if message.text == "Привет":
        await bot.send_message(message.from_user.id, "Привет, для начала работы пиши - /start", reply_markup=markup_start)           
    else:
        await bot.send_message(message.chat.id, "Для начала работы со мной необходимо написать - /start", reply_markup=markup_start)

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
